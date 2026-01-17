"""
Excel ÑĞºÑĞ¿Ğ¾Ñ€Ñ‚ v10.0 â€” HR Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚ Ğ´Ğ»Ñ AI Analyzer v7.0 (Hybrid Expert)

Design: Dark Industrial â€” Ğ¼Ğ¾Ğ½Ğ¾Ñ…Ñ€Ğ¾Ğ¼ Ñ Ğ°ĞºÑ†ĞµĞ½Ñ‚Ğ°Ğ¼Ğ¸
Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚: Ğ’ĞµÑ€Ğ´Ğ¸ĞºÑ‚Ñ‹ High/Medium/Low/Mismatch + Must-haves + Reasoning

2 Ğ²ĞºĞ»Ğ°Ğ´ĞºĞ¸:
- Ğ¨Ğ¾Ñ€Ñ‚Ğ»Ğ¸ÑÑ‚: Ğ‘Ñ‹ÑÑ‚Ñ€Ñ‹Ğ¹ Ğ¾Ğ±Ğ·Ğ¾Ñ€ ĞºĞ°Ğ½Ğ´Ğ¸Ğ´Ğ°Ñ‚Ğ¾Ğ² Ğ·Ğ° 30 ÑĞµĞºÑƒĞ½Ğ´
- ĞŸĞ¾Ğ»Ğ½Ñ‹Ğ¹ Ğ°Ğ½Ğ°Ğ»Ğ¸Ğ·: reasoning_for_hr, must-haves, Ğ²Ğ¾Ğ¿Ñ€Ğ¾ÑÑ‹ Ñ checks
"""
from fastapi.responses import FileResponse
from typing import Optional
from datetime import datetime
import tempfile
import json


def create_excel_export(vacancy, results: list, recommendation_filter: Optional[str] = None) -> str:
    """Excel v10.0 â€” Ğ¿Ğ¾Ğ´ AI v7.0 Hybrid Expert"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    shortlist_ws = wb.create_sheet("Ğ¨Ğ¾Ñ€Ñ‚Ğ»Ğ¸ÑÑ‚", 0)
    deep_ws = wb.create_sheet("ĞŸĞ¾Ğ»Ğ½Ñ‹Ğ¹ Ğ°Ğ½Ğ°Ğ»Ğ¸Ğ·", 1)

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # DESIGN SYSTEM â€” Dark Industrial
    # ĞœĞ¾Ğ½Ğ¾Ñ…Ñ€Ğ¾Ğ¼Ğ½Ğ°Ñ Ğ±Ğ°Ğ·Ğ° (zinc) + ÑÑ€ĞºĞ¸Ğµ Ğ°ĞºÑ†ĞµĞ½Ñ‚Ñ‹ Ğ´Ğ»Ñ Ğ²ĞµÑ€Ğ´Ğ¸ĞºÑ‚Ğ¾Ğ²
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    PALETTE = {
        # Base â€” Dark Industrial
        'ink': '18181B',       # zinc-900 â€” Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸
        'charcoal': '27272A',  # zinc-800 â€” header bg
        'steel': '3F3F46',     # zinc-700 â€” secondary
        'slate': '71717A',     # zinc-500 â€” muted text
        'silver': 'A1A1AA',    # zinc-400 â€” subtle
        'cloud': 'F4F4F5',     # zinc-100 â€” row alt
        'white': 'FFFFFF',
        'border': 'E4E4E7',    # zinc-200

        # Verdict accents â€” Bold & Clear
        'emerald': '059669',      # High â€” Ñ€ĞµĞºĞ¾Ğ¼ĞµĞ½Ğ´ÑƒÑ
        'emerald_bg': 'D1FAE5',
        'blue': '2563EB',         # Medium â€” Ğ½Ğ° Ñ€Ğ°ÑÑĞ¼Ğ¾Ñ‚Ñ€ĞµĞ½Ğ¸Ğµ
        'blue_bg': 'DBEAFE',
        'amber': 'D97706',        # Low â€” ÑĞ¾Ğ¼Ğ½Ğ¸Ñ‚ĞµĞ»ÑŒĞ½Ğ¾
        'amber_bg': 'FEF3C7',
        'red': 'DC2626',          # Mismatch â€” Ğ½Ğµ Ğ¿Ğ¾Ğ´Ñ…Ğ¾Ğ´Ğ¸Ñ‚
        'red_bg': 'FEE2E2',
    }

    # Typography â€” Clean & Professional
    FONT = {
        'title': Font(name='Segoe UI Semibold', size=16, bold=True, color=PALETTE['ink']),
        'subtitle': Font(name='Segoe UI', size=10, color=PALETTE['slate']),
        'header': Font(name='Segoe UI Semibold', size=9, bold=True, color=PALETTE['white']),
        'name': Font(name='Segoe UI Semibold', size=11, bold=True, color=PALETTE['ink']),
        'body': Font(name='Segoe UI', size=10, color=PALETTE['steel']),
        'small': Font(name='Segoe UI', size=9, color=PALETTE['slate']),
        'link': Font(name='Segoe UI', size=10, color=PALETTE['blue'], underline='single'),
        'reasoning': Font(name='Segoe UI', size=10, color=PALETTE['ink']),
    }

    FILL = {
        'header': PatternFill(start_color=PALETTE['charcoal'], fill_type="solid"),
        'cloud': PatternFill(start_color=PALETTE['cloud'], fill_type="solid"),
        'white': PatternFill(start_color=PALETTE['white'], fill_type="solid"),
        # Verdict backgrounds
        'high': PatternFill(start_color=PALETTE['emerald_bg'], fill_type="solid"),
        'medium': PatternFill(start_color=PALETTE['blue_bg'], fill_type="solid"),
        'low': PatternFill(start_color=PALETTE['amber_bg'], fill_type="solid"),
        'mismatch': PatternFill(start_color=PALETTE['red_bg'], fill_type="solid"),
    }

    BORDER = Border(
        left=Side(style='thin', color=PALETTE['border']),
        right=Side(style='thin', color=PALETTE['border']),
        top=Side(style='thin', color=PALETTE['border']),
        bottom=Side(style='thin', color=PALETTE['border'])
    )

    ALIGN = {
        'center': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'left': Alignment(horizontal="left", vertical="center", wrap_text=True),
        'top': Alignment(horizontal="left", vertical="top", wrap_text=True),
    }

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # HELPERS
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    def raw(analysis, key, default=None):
        """ĞŸĞ¾Ğ»ÑƒÑ‡Ğ¸Ñ‚ÑŒ Ğ·Ğ½Ğ°Ñ‡ĞµĞ½Ğ¸Ğµ Ğ¸Ğ· raw_result"""
        r = getattr(analysis, 'raw_result', None) or {}
        return r.get(key, default)

    def get_verdict(analysis):
        """ĞŸĞ¾Ğ»ÑƒÑ‡Ğ¸Ñ‚ÑŒ Ğ²ĞµÑ€Ğ´Ğ¸ĞºÑ‚ v7.0 Ğ¸Ğ»Ğ¸ ÑĞ¼Ğ°Ğ¿Ğ¸Ñ‚ÑŒ Ğ¸Ğ· ÑÑ‚Ğ°Ñ€Ğ¾Ğ³Ğ¾ Ñ„Ğ¾Ñ€Ğ¼Ğ°Ñ‚Ğ°"""
        v = raw(analysis, 'verdict')
        if v in ['High', 'Medium', 'Low', 'Mismatch']:
            return v
        # ĞœĞ°Ğ¿Ğ¿Ğ¸Ğ½Ğ³ ÑÑ‚Ğ°Ñ€Ñ‹Ñ… Ğ·Ğ½Ğ°Ñ‡ĞµĞ½Ğ¸Ğ¹
        old_v = raw(analysis, 'verdict', '')
        if old_v == 'GREEN':
            return 'High'
        elif old_v == 'YELLOW':
            return 'Medium'
        elif old_v == 'RED':
            return 'Low'
        # ĞŸĞ¾ score ĞµÑĞ»Ğ¸ Ğ½ĞµÑ‚ verdict
        score = raw(analysis, 'score', 0) or 0
        if score >= 75:
            return 'High'
        elif score >= 50:
            return 'Medium'
        elif score >= 25:
            return 'Low'
        return 'Mismatch'

    def verdict_display(v):
        """Ğ’ĞµÑ€Ğ´Ğ¸ĞºÑ‚ â†’ (label_ru, fill, color, icon)"""
        return {
            'High': ('Ğ ĞµĞºĞ¾Ğ¼ĞµĞ½Ğ´ÑƒÑ', FILL['high'], PALETTE['emerald'], 'â˜…'),
            'Medium': ('ĞĞ° Ñ€Ğ°ÑÑĞ¼Ğ¾Ñ‚Ñ€ĞµĞ½Ğ¸Ğµ', FILL['medium'], PALETTE['blue'], 'â—†'),
            'Low': ('Ğ¡Ğ¾Ğ¼Ğ½Ğ¸Ñ‚ĞµĞ»ÑŒĞ½Ğ¾', FILL['low'], PALETTE['amber'], 'â–²'),
            'Mismatch': ('ĞĞµ Ğ¿Ğ¾Ğ´Ñ…Ğ¾Ğ´Ğ¸Ñ‚', FILL['mismatch'], PALETTE['red'], 'âœ•'),
        }.get(v, ('â€”', FILL['white'], PALETTE['slate'], '?'))

    def get_priority(analysis):
        """ĞŸĞ¾Ğ»ÑƒÑ‡Ğ¸Ñ‚ÑŒ Ğ¿Ñ€Ğ¸Ğ¾Ñ€Ğ¸Ñ‚ĞµÑ‚ v7.2"""
        p = raw(analysis, 'priority', 'basic')
        return p if p in ['top', 'strong', 'basic'] else 'basic'

    def priority_stars(p):
        """ĞŸÑ€Ğ¸Ğ¾Ñ€Ğ¸Ñ‚ĞµÑ‚ â†’ Ğ·Ğ²Ñ‘Ğ·Ğ´Ñ‹"""
        return {'top': 'â˜…â˜…â˜…', 'strong': 'â˜…â˜…', 'basic': 'â˜…'}.get(p, 'â˜…')

    def get_one_liner(analysis):
        """ĞŸĞ¾Ğ»ÑƒÑ‡Ğ¸Ñ‚ÑŒ one_liner Ğ¸Ğ»Ğ¸ fallback"""
        one = raw(analysis, 'one_liner', '')
        if one:
            return one
        # Fallback to verdict_reason
        return raw(analysis, 'verdict_reason', '') or ''

    def get_salary_from_resume(app):
        """ĞŸĞ¾Ğ»ÑƒÑ‡Ğ¸Ñ‚ÑŒ Ğ·Ğ°Ñ€Ğ¿Ğ»Ğ°Ñ‚Ñƒ Ğ¸Ğ· Ñ€ĞµĞ·ÑĞ¼Ğµ ĞºĞ°Ğ½Ğ´Ğ¸Ğ´Ğ°Ñ‚Ğ°"""
        resume = app.resume_data or {}
        if isinstance(resume, str):
            try:
                resume = json.loads(resume)
            except:
                return None
        salary = resume.get('salary', {})
        if isinstance(salary, dict):
            amount = salary.get('amount', 0) or salary.get('from', 0) or 0
            if amount:
                return int(amount * 1.15)  # NET â†’ GROSS
        return None

    def has_cover_letter(app):
        """ĞŸÑ€Ğ¾Ğ²ĞµÑ€Ğ¸Ñ‚ÑŒ Ğ½Ğ°Ğ»Ğ¸Ñ‡Ğ¸Ğµ ÑĞ¾Ğ¿Ñ€Ğ¾Ğ²Ğ¾Ğ´Ğ¸Ñ‚ĞµĞ»ÑŒĞ½Ğ¾Ğ³Ğ¾ Ğ¿Ğ¸ÑÑŒĞ¼Ğ°"""
        resume = app.resume_data or {}
        if isinstance(resume, str):
            try:
                resume = json.loads(resume)
            except:
                return False
        cover = resume.get('cover_letter', '') or resume.get('message', '')
        return bool(cover and len(str(cover).strip()) > 10)

    def bullets(items, max_n=5):
        if not items:
            return "â€”"
        if isinstance(items, str):
            return items
        return "\n".join([f"â€¢ {i}" for i in items[:max_n] if i])

    def cell_style(cell, font=None, fill=None, align=None, border=None):
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        if border:
            cell.border = border

    def format_must_haves_v7(must_haves):
        """Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚Ğ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ¸Ğµ must-haves v7.0 (yes/maybe/no)"""
        if not must_haves:
            return "â€”"
        lines = []
        for m in must_haves[:5]:
            if isinstance(m, dict):
                req = m.get('requirement', '')
                status = m.get('status', 'no')
                icon = 'âœ“' if status == 'yes' else ('?' if status == 'maybe' else 'âœ—')
                evidence = m.get('evidence', '') or ''
                line = f"{icon} {req}"
                if evidence:
                    line += f"\n   â†’ {evidence}"
                lines.append(line)
        return "\n".join(lines) if lines else "â€”"

    def format_interview_questions_v7(questions):
        """Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚Ğ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ¸Ğµ Ğ²Ğ¾Ğ¿Ñ€Ğ¾ÑĞ¾Ğ² v7.0 (question + checks)"""
        if not questions:
            return "â€”"
        lines = []
        for i, q in enumerate(questions[:4], 1):
            if isinstance(q, dict):
                question = q.get('question', '')
                checks = q.get('checks', '')
                lines.append(f"{i}. {question}")
                if checks:
                    lines.append(f"   ĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑĞµĞ¼: {checks}")
                lines.append("")
            elif isinstance(q, str):
                lines.append(f"{i}. {q}")
        return "\n".join(lines).strip() if lines else "â€”"

    def format_growth_pattern(pattern):
        """Ğ¢Ñ€Ğ°ĞµĞºÑ‚Ğ¾Ñ€Ğ¸Ñ ĞºĞ°Ñ€ÑŒĞµÑ€Ñ‹"""
        return {
            'Ñ€Ğ°ÑÑ‚Ñ‘Ñ‚': 'â†— Ğ Ğ°ÑÑ‚Ñ‘Ñ‚',
            'ÑÑ‚Ğ°Ğ±Ğ¸Ğ»ĞµĞ½': 'â†’ Ğ¡Ñ‚Ğ°Ğ±Ğ¸Ğ»ĞµĞ½',
            'Ğ´ĞµĞ³Ñ€Ğ°Ğ´Ğ¸Ñ€ÑƒĞµÑ‚': 'â†˜ Ğ”ĞµĞ³Ñ€Ğ°Ğ´Ğ¸Ñ€ÑƒĞµÑ‚',
            'Ğ½ĞµĞ¿Ğ¾Ğ½ÑÑ‚Ğ½Ğ¾': '? ĞĞµĞ¿Ğ¾Ğ½ÑÑ‚Ğ½Ğ¾',
        }.get(pattern, 'â€”')

    def format_salary_fit(salary_fit):
        """Ğ¡Ğ¾Ğ¾Ñ‚Ğ²ĞµÑ‚ÑÑ‚Ğ²Ğ¸Ğµ Ğ·Ğ°Ñ€Ğ¿Ğ»Ğ°Ñ‚Ñ‹"""
        if not salary_fit:
            return "â€”"
        if isinstance(salary_fit, dict):
            status = salary_fit.get('status', '')
            comment = salary_fit.get('comment', '')
            return f"{status}" + (f" ({comment})" if comment else "")
        return str(salary_fit)

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # STATISTICS
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    total = len(results)
    high_n = len([r for r, _ in results if get_verdict(r) == 'High'])
    medium_n = len([r for r, _ in results if get_verdict(r) == 'Medium'])
    low_n = len([r for r, _ in results if get_verdict(r) == 'Low'])
    mismatch_n = total - high_n - medium_n - low_n

    # Ğ¡Ğ¾Ñ€Ñ‚Ğ¸Ñ€Ğ¾Ğ²ĞºĞ°: High+top â†’ High+strong â†’ High+basic â†’ Medium â†’ Low â†’ Mismatch
    def sort_key(item):
        v = get_verdict(item[0])
        p = get_priority(item[0])
        verdict_order = {'High': 0, 'Medium': 1, 'Low': 2, 'Mismatch': 3}
        priority_order = {'top': 0, 'strong': 1, 'basic': 2}
        score = raw(item[0], 'score', 0) or 0
        return (verdict_order.get(v, 3), priority_order.get(p, 2), -score)

    sorted_results = sorted(results, key=sort_key)

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # SHEET 1: Ğ¨ĞĞ Ğ¢Ğ›Ğ˜Ğ¡Ğ¢ â€” Ğ‘Ñ‹ÑÑ‚Ñ€Ñ‹Ğ¹ Ğ¾Ğ±Ğ·Ğ¾Ñ€ Ğ·Ğ° 30 ÑĞµĞºÑƒĞ½Ğ´
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    ws = shortlist_ws

    # Column widths v7.2: ĞŸÑ€Ğ¸Ğ¾Ñ€Ğ¸Ñ‚ĞµÑ‚ | ĞšĞ°Ğ½Ğ´Ğ¸Ğ´Ğ°Ñ‚ | Ğ’ĞµÑ€Ğ´Ğ¸ĞºÑ‚ | Ğ—Ğ°Ñ€Ğ¿Ğ»Ğ°Ñ‚Ğ° | ĞŸĞ¸ÑÑŒĞ¼Ğ¾ | ĞŸĞ¾Ñ‡ĞµĞ¼Ñƒ ÑÑ‚Ğ¾Ñ‚ ĞºĞ°Ğ½Ğ´Ğ¸Ğ´Ğ°Ñ‚
    widths = {'A': 6, 'B': 22, 'C': 14, 'D': 14, 'E': 7, 'F': 70}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title row
    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value=vacancy.title)
    cell_style(c, font=FONT['title'], align=ALIGN['center'])
    ws.row_dimensions[1].height = 36

    # Stats bar v7.2
    ws.merge_cells('A2:F2')
    # Count top priority
    top_count = len([r for r, _ in results if get_verdict(r) == 'High' and get_priority(r) == 'top'])
    stats = f"â˜…â˜…â˜… {top_count} Ñ‚Ğ¾Ğ¿   â”‚   â˜… {high_n} Ñ€ĞµĞºĞ¾Ğ¼ĞµĞ½Ğ´ÑƒÑ   â—† {medium_n} Ñ€Ğ°ÑÑĞ¼Ğ¾Ñ‚Ñ€ĞµÑ‚ÑŒ   â–² {low_n} ÑĞ¾Ğ¼Ğ½Ğ¸Ñ‚ĞµĞ»ÑŒĞ½Ğ¾   âœ• {mismatch_n} Ğ½ĞµÑ‚   â”‚   {total} Ğ²ÑĞµĞ³Ğ¾"
    c = ws.cell(row=2, column=1, value=stats)
    cell_style(c, font=FONT['subtitle'], align=ALIGN['center'], fill=FILL['cloud'])
    ws.row_dimensions[2].height = 26

    # Spacer
    ws.row_dimensions[3].height = 6

    # Headers v7.2
    headers = ['â˜…', 'ĞšĞĞĞ”Ğ˜Ğ”ĞĞ¢', 'Ğ’Ğ•Ğ Ğ”Ğ˜ĞšĞ¢', 'Ğ—ĞĞ ĞŸĞ›ĞĞ¢Ğ', 'ğŸ“', 'ĞŸĞĞ§Ğ•ĞœĞ£ Ğ­Ğ¢ĞĞ¢ ĞšĞĞĞ”Ğ˜Ğ”ĞĞ¢']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        cell_style(c, font=FONT['header'], fill=FILL['header'], align=ALIGN['center'], border=BORDER)
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = 'A5'

    # Data rows v7.2
    for idx, (analysis, app) in enumerate(sorted_results, 1):
        row = idx + 4
        verdict = get_verdict(analysis)
        priority = get_priority(analysis)
        label_ru, fill_v, color_v, icon = verdict_display(verdict)
        row_fill = FILL['white'] if idx % 2 else FILL['cloud']

        # Col 1: Priority stars (Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ High)
        if verdict == 'High':
            stars = priority_stars(priority)
            c = ws.cell(row=row, column=1, value=stars)
            cell_style(c, font=Font(name='Segoe UI', size=11, color=PALETTE['amber']),
                       fill=row_fill, align=ALIGN['center'], border=BORDER)
        else:
            c = ws.cell(row=row, column=1, value=icon)
            cell_style(c, font=Font(name='Segoe UI', size=12, color=color_v),
                       fill=fill_v, align=ALIGN['center'], border=BORDER)

        # Col 2: Name (with link)
        name = app.candidate_name or "â€”"
        c = ws.cell(row=row, column=2, value=name)
        if app.resume_url:
            c.hyperlink = app.resume_url
            cell_style(c, font=FONT['link'], fill=row_fill, align=ALIGN['left'], border=BORDER)
        else:
            cell_style(c, font=FONT['name'], fill=row_fill, align=ALIGN['left'], border=BORDER)

        # Col 3: Verdict label
        c = ws.cell(row=row, column=3, value=label_ru)
        cell_style(c, font=Font(name='Segoe UI Semibold', size=10, bold=True, color=color_v),
                   fill=fill_v, align=ALIGN['center'], border=BORDER)

        # Col 4: Salary from resume
        salary = get_salary_from_resume(app)
        if salary:
            salary_text = f"{salary:,}".replace(',', ' ') + " â‚½"
        else:
            salary_text = "â€”"
        c = ws.cell(row=row, column=4, value=salary_text)
        cell_style(c, font=FONT['body'], fill=row_fill, align=ALIGN['center'], border=BORDER)

        # Col 5: Cover letter indicator
        has_cover = has_cover_letter(app)
        c = ws.cell(row=row, column=5, value="âœ“" if has_cover else "â€”")
        cover_color = PALETTE['emerald'] if has_cover else PALETTE['slate']
        cell_style(c, font=Font(name='Segoe UI', size=10, color=cover_color),
                   fill=row_fill, align=ALIGN['center'], border=BORDER)

        # Col 6: One-liner (Ğ¿Ğ¾Ñ‡ĞµĞ¼Ñƒ ÑÑ‚Ğ¾Ñ‚ ĞºĞ°Ğ½Ğ´Ğ¸Ğ´Ğ°Ñ‚)
        one_liner = get_one_liner(analysis)
        c = ws.cell(row=row, column=6, value=one_liner or "â€”")
        cell_style(c, font=FONT['reasoning'], fill=row_fill, align=ALIGN['top'], border=BORDER)

        ws.row_dimensions[row].height = 52

    ws.auto_filter.ref = f"A4:F{len(results) + 4}"

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # SHEET 2: ĞŸĞĞ›ĞĞ«Ğ™ ĞĞĞĞ›Ğ˜Ğ— â€” v7.0 Hybrid Expert
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    ws = deep_ws

    # Columns: â„– | ĞšĞĞĞ”Ğ˜Ğ”ĞĞ¢ | Ğ’Ğ•Ğ Ğ”Ğ˜ĞšĞ¢ | MUST-HAVES | Ğ Ğ•Ğ—Ğ®ĞœĞ• ĞšĞĞ Ğ¬Ğ•Ğ Ğ« | ĞŸĞ›Ğ®Ğ¡Ğ«/ĞœĞ˜ĞĞ£Ğ¡Ğ« | Ğ’ĞĞŸĞ ĞĞ¡Ğ«
    widths = {'A': 4, 'B': 18, 'C': 14, 'D': 36, 'E': 42, 'F': 32, 'G': 48}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1, value=f"ĞŸĞ¾Ğ»Ğ½Ñ‹Ğ¹ Ğ°Ğ½Ğ°Ğ»Ğ¸Ğ·: {vacancy.title}")
    cell_style(c, font=FONT['title'], align=ALIGN['center'])
    ws.row_dimensions[1].height = 36

    # Spacer
    ws.row_dimensions[2].height = 6

    # Headers
    headers = ['â„–', 'ĞšĞĞĞ”Ğ˜Ğ”ĞĞ¢', 'Ğ’Ğ•Ğ Ğ”Ğ˜ĞšĞ¢', 'MUST-HAVES', 'ĞĞĞĞ›Ğ˜Ğ— ĞšĞĞ Ğ¬Ğ•Ğ Ğ«', 'ĞĞ¦Ğ•ĞĞšĞ', 'Ğ’ĞĞŸĞ ĞĞ¡Ğ« Ğ”Ğ›Ğ¯ Ğ˜ĞĞ¢Ğ•Ğ Ğ’Ğ¬Ğ®']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        cell_style(c, font=FONT['header'], fill=FILL['header'], align=ALIGN['center'], border=BORDER)
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = 'C4'

    # Data
    for idx, (analysis, app) in enumerate(sorted_results, 1):
        row = idx + 3
        verdict = get_verdict(analysis)
        label_ru, fill_v, color_v, icon = verdict_display(verdict)
        row_fill = FILL['white'] if idx % 2 else FILL['cloud']

        # Col 1: â„–
        c = ws.cell(row=row, column=1, value=idx)
        cell_style(c, font=FONT['small'], fill=row_fill, align=ALIGN['center'], border=BORDER)

        # Col 2: Name (with link)
        name = app.candidate_name or "â€”"
        c = ws.cell(row=row, column=2, value=name)
        if app.resume_url:
            c.hyperlink = app.resume_url
            cell_style(c, font=FONT['link'], fill=row_fill, align=ALIGN['left'], border=BORDER)
        else:
            cell_style(c, font=FONT['name'], fill=row_fill, align=ALIGN['left'], border=BORDER)

        # Col 3: Verdict
        verdict_text = f"{icon} {label_ru}"
        c = ws.cell(row=row, column=3, value=verdict_text)
        cell_style(c, font=Font(name='Segoe UI Semibold', size=10, bold=True, color=color_v),
                   fill=fill_v, align=ALIGN['center'], border=BORDER)

        # Col 4: Must-haves v7.0
        must_haves = raw(analysis, 'must_haves', []) or raw(analysis, 'must_have', [])
        c = ws.cell(row=row, column=4, value=format_must_haves_v7(must_haves))
        # Color by status
        no_count = sum(1 for m in (must_haves or []) if isinstance(m, dict) and m.get('status') == 'no')
        mh_color = PALETTE['red'] if no_count > 0 else PALETTE['emerald']
        cell_style(c, font=Font(name='Segoe UI', size=9, color=mh_color),
                   fill=row_fill, align=ALIGN['top'], border=BORDER)

        # Col 5: Career analysis (holistic + reasoning)
        holistic = raw(analysis, 'holistic_analysis', {}) or {}
        career_summary = holistic.get('career_summary', '') if isinstance(holistic, dict) else ''
        relevance = holistic.get('relevance_assessment', '') if isinstance(holistic, dict) else ''
        growth = holistic.get('growth_pattern', '') if isinstance(holistic, dict) else ''
        reasoning = raw(analysis, 'reasoning_for_hr', '') or ''

        career_text_parts = []
        if career_summary:
            career_text_parts.append(f"ĞšĞ°Ñ€ÑŒĞµÑ€Ğ°: {career_summary}")
        if relevance:
            career_text_parts.append(f"Ğ ĞµĞ»ĞµĞ²Ğ°Ğ½Ñ‚Ğ½Ğ¾ÑÑ‚ÑŒ: {relevance}")
        if growth:
            career_text_parts.append(f"Ğ¢Ñ€Ğ°ĞµĞºÑ‚Ğ¾Ñ€Ğ¸Ñ: {format_growth_pattern(growth)}")
        if reasoning:
            career_text_parts.append(f"\n{reasoning}")

        career_text = "\n".join(career_text_parts) if career_text_parts else "â€”"
        c = ws.cell(row=row, column=5, value=career_text)
        cell_style(c, font=FONT['body'], fill=row_fill, align=ALIGN['top'], border=BORDER)

        # Col 6: Pros/Cons + Concerns + Salary
        concerns = raw(analysis, 'concerns', []) or []
        salary_fit = raw(analysis, 'salary_fit', None)
        strengths = raw(analysis, 'strengths', []) or raw(analysis, 'pros', []) or []
        weaknesses = raw(analysis, 'weaknesses', []) or raw(analysis, 'cons', []) or []

        assessment_parts = []
        if strengths:
            assessment_parts.append("ĞŸĞ»ÑÑÑ‹:\n" + bullets(strengths, 3))
        if weaknesses or concerns:
            all_concerns = list(weaknesses) + list(concerns)
            assessment_parts.append("ĞœĞ¸Ğ½ÑƒÑÑ‹:\n" + bullets(all_concerns, 3))
        if salary_fit:
            assessment_parts.append(f"Ğ—Ğ°Ñ€Ğ¿Ğ»Ğ°Ñ‚Ğ°: {format_salary_fit(salary_fit)}")

        c = ws.cell(row=row, column=6, value="\n\n".join(assessment_parts) if assessment_parts else "â€”")
        cell_style(c, font=FONT['body'], fill=row_fill, align=ALIGN['top'], border=BORDER)

        # Col 7: Interview Questions v7.0 (with checks)
        questions = raw(analysis, 'interview_questions_v7', []) or raw(analysis, 'interview_questions', [])
        c = ws.cell(row=row, column=7, value=format_interview_questions_v7(questions))
        cell_style(c, font=Font(name='Segoe UI', size=9, color=PALETTE['ink']),
                   fill=row_fill, align=ALIGN['top'], border=BORDER)

        ws.row_dimensions[row].height = 130

    ws.auto_filter.ref = f"A3:G{len(results) + 3}"

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # SAVE
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp.name)
    temp.close()
    return temp.name
