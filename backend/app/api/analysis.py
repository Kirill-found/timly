"""
API endpoints для AI анализа резюме
Запуск анализа, получение результатов, экспорт
"""
from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional

from app.database import get_db
from app.api.auth import get_current_user
from app.schemas.base import APIResponse
from app.schemas.analysis import (
    AnalysisRequest, AnalysisResult as AnalysisResultSchema, BatchAnalysisResponse,
    AnalysisFilter, AnalysisStats, ExportRequest
)
from app.models.vacancy import Vacancy
from app.models.application import Application, AnalysisResult
from app.services.ai_analyzer import AIAnalyzer
from app.utils.exceptions import AIAnalysisError, ValidationError
from app.utils.response import success, created, bad_request, unauthorized, not_found, internal_error
from app.utils.logger import get_logger

logger = get_logger(__name__)

router = APIRouter()


@router.post("/start", response_model=APIResponse)
async def start_analysis(
    analysis_request: AnalysisRequest,
    background_tasks: BackgroundTasks,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Запуск AI анализа резюме
    Максимум 5 резюме за раз для оптимизации стоимости
    """
    try:
        from app.models.application import Application
        from app.workers.analysis_jobs import run_ai_analysis_batch
        import uuid

        applications_count = len(analysis_request.application_ids)

        # Проверка существования откликов и принадлежности пользователю
        applications = db.query(Application).join(Application.vacancy).filter(
            Application.id.in_(analysis_request.application_ids),
            Application.vacancy.has(user_id=current_user.id)
        ).all()

        if len(applications) != applications_count:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={
                    "error": "APPLICATIONS_NOT_FOUND",
                    "message": "Некоторые отклики не найдены"
                }
            )

        job_id = str(uuid.uuid4())

        # Запуск анализа в фоновой задаче (без передачи db - создаётся внутри)
        background_tasks.add_task(
            run_ai_analysis_batch,
            analysis_request.application_ids,
            current_user.id,
            analysis_request.force_reanalysis
        )

        analysis_data = {
            "job_id": job_id,
            "message": "Анализ поставлен в очередь",
            "applications_queued": applications_count,
            "estimated_time_seconds": applications_count * 15  # 15 сек на резюме
        }

        return success(data=analysis_data)
    except ValidationError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error": "VALIDATION_ERROR",
                "message": str(e)
            }
        )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Ошибка при запуске анализа: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error": "ANALYSIS_START_ERROR",
                "message": f"Ошибка при запуске анализа: {str(e)}"
            }
        )


@router.post("/start-new", response_model=APIResponse)
async def start_analysis_new_applications(
    vacancy_id: str,
    collection_filter: Optional[str] = None,
    limit: Optional[int] = None,
    background_tasks: BackgroundTasks = BackgroundTasks(),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Запуск анализа ТОЛЬКО новых (непроанализированных) откликов

    По умолчанию анализирует отклики 'response' (новые) и 'consider' (рассматриваемые),
    но НЕ анализирует 'discard' (отклоненные)
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[START-NEW] Запрос на анализ новых откликов: vacancy_id={vacancy_id}, user={current_user.id}")

    try:
        from app.models.application import Application, AnalysisResult
        from app.models.vacancy import Vacancy
        from app.models.subscription import Subscription, SubscriptionStatus
        from app.workers.analysis_jobs import run_ai_analysis_batch
        import uuid

        # ПРОВЕРКА ЛИМИТОВ ПОДПИСКИ
        active_subscription = db.query(Subscription).filter(
            Subscription.user_id == current_user.id,
            Subscription.status.in_([SubscriptionStatus.active, SubscriptionStatus.trial])
        ).first()

        if not active_subscription:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={
                    "error": "NO_SUBSCRIPTION",
                    "message": "У вас нет активной подписки. Пожалуйста, выберите тариф."
                }
            )

        # Проверка возможности анализа (лимиты)
        can_analyze, error_message = active_subscription.can_analyze()
        if not can_analyze:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={
                    "error": "LIMIT_EXCEEDED",
                    "message": error_message,
                    "subscription": {
                        "plan_type": active_subscription.plan.plan_type.value if active_subscription.plan else "unknown",
                        "analyses_used": active_subscription.analyses_used_this_month,
                        "analyses_limit": active_subscription.plan.max_analyses_per_month if active_subscription.plan else 0
                    }
                }
            )

        # Проверка существования вакансии и принадлежности пользователю
        vacancy = db.query(Vacancy).filter(
            Vacancy.id == vacancy_id,
            Vacancy.user_id == current_user.id
        ).first()

        if not vacancy:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={
                    "error": "VACANCY_NOT_FOUND",
                    "message": "Вакансия не найдена"
                }
            )

        # Найти все непроанализированные отклики
        query = db.query(Application).outerjoin(
            AnalysisResult,
            Application.id == AnalysisResult.application_id
        ).filter(
            Application.vacancy_id == vacancy_id,
            AnalysisResult.id == None  # Нет анализа
        )

        # Фильтр по collection_id
        if collection_filter:
            # Если указан конкретный фильтр - используем его
            query = query.filter(Application.collection_id.like(f"%{collection_filter}%"))
        else:
            # По умолчанию: анализируем response и consider, НО НЕ discard
            from sqlalchemy import or_, and_, not_
            query = query.filter(
                and_(
                    or_(
                        Application.collection_id.like("%response%"),
                        Application.collection_id.like("%consider%")
                    ),
                    not_(Application.collection_id.like("%discard%"))
                )
            )

        # Применяем limit если указан
        if limit and limit > 0:
            query = query.limit(limit)

        unanalyzed_applications = query.all()
        application_ids = [str(app.id) for app in unanalyzed_applications]

        logger.info(f"[START-NEW] Найдено {len(application_ids)} непроанализированных откликов (limit={limit if limit else 'не указан'})")

        if not application_ids:
            logger.info("[START-NEW] Нет откликов для анализа")
            return success(data={
                "message": "Нет новых откликов для анализа",
                "applications_queued": 0
            })

        job_id = str(uuid.uuid4())
        user_id_str = str(current_user.id)

        logger.info(f"[START-NEW] Запуск фоновой задачи: job_id={job_id}, apps={len(application_ids)}, user={user_id_str}")

        # Запуск анализа в фоновой задаче
        background_tasks.add_task(
            run_ai_analysis_batch,
            application_ids,
            user_id_str,  # Преобразуем UUID в строку
            False  # НЕ принудительный анализ
        )

        logger.info(f"[START-NEW] Фоновая задача добавлена в очередь")

        analysis_data = {
            "job_id": job_id,
            "message": f"Анализ {len(application_ids)} новых откликов поставлен в очередь",
            "applications_queued": len(application_ids),
            "collection_filter": collection_filter,
            "estimated_time_seconds": len(application_ids) * 15
        }

        return success(data=analysis_data)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Ошибка при запуске анализа новых откликов: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error": "ANALYSIS_START_ERROR",
                "message": f"Ошибка при запуске анализа: {str(e)}"
            }
        )


@router.post("/reanalyze-all", response_model=APIResponse)
async def reanalyze_all_applications(
    vacancy_id: str,
    background_tasks: BackgroundTasks = BackgroundTasks(),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Переанализ ВСЕХ проанализированных откликов с новым промптом

    Использует force_reanalysis=True для обновления существующих анализов
    Полезно после обновления промпта AI для получения единой методологии оценки
    """
    try:
        from app.models.application import Application, AnalysisResult
        from app.models.vacancy import Vacancy
        from app.workers.analysis_jobs import run_ai_analysis_batch
        import uuid

        # Проверка вакансии
        vacancy = db.query(Vacancy).filter(
            Vacancy.id == vacancy_id,
            Vacancy.user_id == current_user.id
        ).first()

        if not vacancy:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={
                    "error": "VACANCY_NOT_FOUND",
                    "message": "Вакансия не найдена"
                }
            )

        # Получаем ВСЕ отклики, которые уже проанализированы
        analyzed_applications = db.query(Application).join(
            AnalysisResult,
            Application.id == AnalysisResult.application_id
        ).filter(
            Application.vacancy_id == vacancy_id
        ).all()

        if not analyzed_applications:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={
                    "error": "NO_ANALYZED_APPLICATIONS",
                    "message": "Нет проанализированных откликов для переанализа"
                }
            )

        application_ids = [str(app.id) for app in analyzed_applications]
        job_id = str(uuid.uuid4())

        # Запуск переанализа в фоновой задаче с force_reanalysis=True
        background_tasks.add_task(
            run_ai_analysis_batch,
            application_ids,
            str(current_user.id),  # Преобразуем UUID в строку
            True  # force_reanalysis=True - ключевой параметр!
        )

        analysis_data = {
            "job_id": job_id,
            "message": f"Переанализ {len(application_ids)} резюме с новым промптом запущен",
            "applications_queued": len(application_ids),
            "estimated_time_seconds": len(application_ids) * 15,
            "force_reanalysis": True
        }

        return success(data=analysis_data)

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Ошибка при запуске переанализа: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error": "REANALYSIS_START_ERROR",
                "message": f"Ошибка при запуске переанализа: {str(e)}"
            }
        )


@router.get("/results", response_model=APIResponse)
async def get_analysis_results(
    filters: AnalysisFilter = Depends(),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Получение результатов AI анализа
    С фильтрацией по различным параметрам
    """
    try:
        from app.models.application import Application, AnalysisResult

        # Базовый запрос: результаты только для откликов пользователя
        query = db.query(AnalysisResult).join(Application).join(Application.vacancy).filter(
            Application.vacancy.has(user_id=current_user.id)
        )

        # Фильтрация по вакансии
        if filters.vacancy_id:
            query = query.filter(Application.vacancy_id == filters.vacancy_id)

        # Фильтрация по минимальному score
        if filters.min_score is not None:
            query = query.filter(AnalysisResult.score >= filters.min_score)

        # Фильтрация по максимальному score
        if filters.max_score is not None:
            query = query.filter(AnalysisResult.score <= filters.max_score)

        # Фильтрация по рекомендации
        if filters.recommendation:
            query = query.filter(AnalysisResult.recommendation == filters.recommendation)

        # Подсчет общего количества
        total = query.count()

        # Применение пагинации и сортировка по score (высокие сначала)
        limit = filters.limit or 50
        offset = filters.offset or 0
        results = query.order_by(AnalysisResult.score.desc().nulls_last()).offset(offset).limit(limit).all()

        # Сериализация с данными кандидата
        results_data = []
        for result in results:
            result_dict = result.to_dict()
            # Добавляем данные кандидата из связанного отклика
            if result.application:
                result_dict['application'] = {
                    'candidate_name': result.application.candidate_name,
                    'candidate_email': result.application.candidate_email,
                    'candidate_phone': result.application.candidate_phone,
                    'resume_url': result.application.resume_url,
                    'created_at': result.application.created_at.isoformat() if result.application.created_at else None,
                }
            results_data.append(result_dict)

        return success(data={
            "results": results_data,
            "total": total,
            "limit": limit,
            "offset": offset,
            "filters_applied": filters.model_dump() if filters else None
        })
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error": "RESULTS_FETCH_ERROR",
                "message": "Ошибка при получении результатов анализа"
            }
        )


@router.get("/results/{analysis_id}", response_model=APIResponse)
async def get_analysis_result(
    analysis_id: str,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Получение конкретного результата анализа
    Детальная информация об одном анализе
    """
    # TODO: Реализовать получение конкретного результата
    return success(data={
        "status": "TODO",
        "message": f"Результат анализа {analysis_id} будет реализован",
        "analysis_id": analysis_id
    })


@router.delete("/results/{analysis_id}", response_model=APIResponse)
async def delete_analysis_result(
    analysis_id: str,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Удаление результата анализа
    Полное удаление из системы
    """
    # TODO: Реализовать удаление результата
    return success(data={
        "status": "TODO",
        "message": f"Удаление результата {analysis_id} будет реализовано",
        "analysis_id": analysis_id,
        "deleted": True
    })


@router.get("/vacancy/{vacancy_id}/stats", response_model=APIResponse)
async def get_vacancy_analysis_stats(
    vacancy_id: str,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Статистика анализа по конкретной вакансии
    Распределение оценок и рекомендаций
    """
    try:
        # Проверяем права доступа к вакансии
        vacancy = db.query(Vacancy).filter(
            Vacancy.id == vacancy_id,
            Vacancy.user_id == current_user.id
        ).first()

        if not vacancy:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={
                    "error": "VACANCY_NOT_FOUND",
                    "message": "Вакансия не найдена"
                }
            )

        # Получаем все результаты анализа для вакансии
        results = db.query(AnalysisResult).join(
            Application, Application.id == AnalysisResult.application_id
        ).filter(
            Application.vacancy_id == vacancy_id
        ).all()

        # Подсчитываем статистику
        total = len(results)
        hire_count = len([r for r in results if r.recommendation == 'hire'])
        interview_count = len([r for r in results if r.recommendation == 'interview'])
        maybe_count = len([r for r in results if r.recommendation == 'maybe'])
        reject_count = len([r for r in results if r.recommendation == 'reject'])

        # Средний балл
        avg_score = 0
        if total > 0:
            total_score = sum([r.score for r in results if r.score is not None])
            avg_score = round(total_score / total, 1) if total > 0 else 0

        # Последний анализ
        last_analysis_at = None
        if results:
            last_result = max(results, key=lambda r: r.created_at)
            last_analysis_at = last_result.created_at.isoformat()

        stats_data = {
            "vacancy_id": vacancy_id,
            "total_analyzed": total,
            "avg_score": avg_score,
            "hire_count": hire_count,
            "interview_count": interview_count,
            "maybe_count": maybe_count,
            "reject_count": reject_count,
            "last_analysis_at": last_analysis_at
        }

        return success(data=stats_data)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting vacancy analysis stats: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error": "STATS_FETCH_ERROR",
                "message": "Ошибка при получении статистики"
            }
        )


@router.get("/export/excel")
async def export_analysis_to_excel(
    vacancy_id: str,
    recommendation: Optional[str] = None,
    min_score: Optional[int] = None,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Экспорт проанализированных откликов в Excel с фильтрацией

    Параметры:
    - vacancy_id: ID вакансии (обязательный)
    - recommendation: Фильтр по рекомендации (hire, interview, maybe, reject)
    - min_score: Минимальный балл для фильтрации
    """
    try:
        from app.models.application import Application, AnalysisResult
        from app.models.vacancy import Vacancy
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        import tempfile
        import os

        # Проверка вакансии
        vacancy = db.query(Vacancy).filter(
            Vacancy.id == vacancy_id,
            Vacancy.user_id == current_user.id
        ).first()

        if not vacancy:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"error": "VACANCY_NOT_FOUND", "message": "Вакансия не найдена"}
            )

        # Базовый запрос результатов с фильтрацией
        query = db.query(AnalysisResult, Application).join(
            Application,
            AnalysisResult.application_id == Application.id
        ).filter(
            Application.vacancy_id == vacancy_id
        )

        # Применение фильтров
        if recommendation:
            query = query.filter(AnalysisResult.recommendation == recommendation)

        if min_score is not None:
            query = query.filter(AnalysisResult.score >= min_score)

        # Сортировка по оценке от большей к меньшей
        results = query.order_by(AnalysisResult.score.desc().nulls_last()).all()

        if not results:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"error": "NO_RESULTS", "message": "Нет проанализированных откликов"}
            )

        # Создание Excel файла
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AI Анализ Резюме"

        # Добавляем лист со статистикой
        stats_ws = wb.create_sheet("Статистика", 0)

        # ========== БРЕНДИРОВАННЫЙ ЗАГОЛОВОК ==========
        # Логотип и название компании (строки 1-3)
        ws.merge_cells('A1:O3')
        logo_cell = ws.cell(row=1, column=1, value="TIMLY\nAI-Powered HR Analytics")
        logo_cell.font = Font(bold=True, size=20, color="FFFFFF", name="Arial")
        logo_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        logo_cell.fill = PatternFill(start_color="6366F1", end_color="8B5CF6", fill_type="solid")  # Градиент синий-фиолетовый
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 25

        # Название вакансии (строка 4)
        ws.merge_cells('A4:O4')
        title_cell = ws.cell(row=4, column=1, value=f"📋 Вакансия: {vacancy.title}")
        title_cell.font = Font(bold=True, size=16, color="1F2937", name="Arial")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        ws.row_dimensions[4].height = 30

        # Информация о дате экспорта (строка 5)
        ws.merge_cells('A5:O5')
        date_cell = ws.cell(row=5, column=1, value=f"📅 Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        date_cell.font = Font(size=11, color="6B7280", italic=True)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 20

        # Заголовок таблицы (строка 7 - после брендинга)
        header = [
            "№", "Кандидат", "Email", "Телефон", "Ссылка на резюме",
            "Оценка", "Навыки", "Опыт", "Зарплата", "Рекомендация",
            "Сильные стороны", "Слабые стороны", "Красные флаги",
            "Обоснование", "Дата анализа"
        ]

        # Современный градиентный стиль заголовков
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12, name="Arial")

        for col, value in enumerate(header, 1):
            cell = ws.cell(row=7, column=col, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[7].height = 30

        # Заморозка панелей (заголовок всегда видим)
        ws.freeze_panes = "A8"

        # Современная цветовая палитра для рекомендаций
        rec_colors = {
            'hire': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),  # Мятно-зеленый
            'interview': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),  # Теплый желтый
            'maybe': PatternFill(start_color="FBCFE8", end_color="FBCFE8", fill_type="solid"),  # Мягкий розовый
            'reject': PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # Нейтральный серый
        }

        rec_font_colors = {
            'hire': "065F46",  # Темно-зеленый
            'interview': "92400E",  # Темно-желтый
            'maybe': "9F1239",  # Темно-розовый
            'reject': "6B7280"  # Темно-серый
        }

        # Перевод рекомендаций на русский с эмодзи
        rec_translations = {
            'hire': '✅ Нанять',
            'interview': '👤 Собеседование',
            'maybe': '🤔 Возможно',
            'reject': '❌ Отклонить'
        }

        # Данные (начинаем с 8-й строки, т.к. строки 1-7 заняты брендингом и заголовками)
        for idx, (analysis, application) in enumerate(results, 8):
            ws.cell(row=idx, column=1, value=idx-7)  # Номер строки (idx-7 т.к. начинаем с 8)
            ws.cell(row=idx, column=2, value=application.candidate_name or "Не указано")
            ws.cell(row=idx, column=3, value=application.candidate_email or "Не указан")
            ws.cell(row=idx, column=4, value=application.candidate_phone or "Не указан")

            # Ссылка на резюме
            if application.resume_url:
                cell = ws.cell(row=idx, column=5, value="Открыть резюме")
                cell.hyperlink = application.resume_url
                cell.font = Font(color="0563C1", underline="single")
            else:
                ws.cell(row=idx, column=5, value="Н/Д")

            # Оценки
            ws.cell(row=idx, column=6, value=analysis.score or 0)
            ws.cell(row=idx, column=7, value=analysis.skills_match or 0)
            ws.cell(row=idx, column=8, value=analysis.experience_match or 0)

            # Зарплата
            salary_text = {
                'match': 'Совпадает',
                'higher': 'Выше ожиданий',
                'lower': 'Ниже ожиданий',
                'unknown': 'Не указана'
            }.get(analysis.salary_match, 'Н/Д')
            ws.cell(row=idx, column=9, value=salary_text)

            # Рекомендация с цветом, эмодзи и переводом на русский
            rec_text = rec_translations.get(analysis.recommendation, analysis.recommendation or "N/A")
            rec_cell = ws.cell(row=idx, column=10, value=rec_text)
            if analysis.recommendation and analysis.recommendation in rec_colors:
                rec_cell.fill = rec_colors[analysis.recommendation]
                rec_cell.font = Font(bold=True, size=11, color=rec_font_colors.get(analysis.recommendation, "000000"))
            rec_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=idx, column=11, value=", ".join(analysis.strengths or []))
            ws.cell(row=idx, column=12, value=", ".join(analysis.weaknesses or []))
            ws.cell(row=idx, column=13, value=", ".join(analysis.red_flags or []))
            ws.cell(row=idx, column=14, value=analysis.reasoning or "")
            ws.cell(row=idx, column=15, value=analysis.created_at.strftime("%Y-%m-%d %H:%M") if analysis.created_at else "N/A")

        # ========== АВТОФИЛЬТР ==========
        # Включаем автофильтр для заголовков (строка 7)
        ws.auto_filter.ref = f"A7:O{len(results) + 7}"

        # ========== УСЛОВНОЕ ФОРМАТИРОВАНИЕ ДЛЯ ОЦЕНОК ==========
        from openpyxl.styles import Color
        from openpyxl.formatting.rule import ColorScaleRule

        # Градиентная раскраска для колонки "Оценка" (F)
        score_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='FFC7CE',  # Красный для низких
            mid_type='num', mid_value=50, mid_color='FFEB9C',  # Желтый для средних
            end_type='num', end_value=100, end_color='C6EFCE'  # Зеленый для высоких
        )
        ws.conditional_formatting.add(f"F8:F{len(results) + 7}", score_rule)

        # Градиентная раскраска для колонки "Навыки" (G)
        skills_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='FFC7CE',
            mid_type='num', mid_value=50, mid_color='FFEB9C',
            end_type='num', end_value=100, end_color='C6EFCE'
        )
        ws.conditional_formatting.add(f"G8:G{len(results) + 7}", skills_rule)

        # Градиентная раскраска для колонки "Опыт" (H)
        exp_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='FFC7CE',
            mid_type='num', mid_value=50, mid_color='FFEB9C',
            end_type='num', end_value=100, end_color='C6EFCE'
        )
        ws.conditional_formatting.add(f"H8:H{len(results) + 7}", exp_rule)

        # ========== АВТОШИРИНА КОЛОНОК ==========
        from openpyxl.cell.cell import MergedCell
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if isinstance(cell, MergedCell):
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # ========== СТРАНИЦА СТАТИСТИКИ ==========
        # Брендированный заголовок
        stats_ws.merge_cells('A1:E3')
        stats_logo = stats_ws.cell(row=1, column=1, value="TIMLY\n📊 Статистика анализа")
        stats_logo.font = Font(bold=True, size=18, color="FFFFFF", name="Arial")
        stats_logo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        stats_logo.fill = PatternFill(start_color="8B5CF6", end_color="6366F1", fill_type="solid")
        stats_ws.row_dimensions[1].height = 20
        stats_ws.row_dimensions[2].height = 20
        stats_ws.row_dimensions[3].height = 20

        # Название вакансии
        stats_ws.merge_cells('A4:E4')
        stats_title = stats_ws.cell(row=4, column=1, value=f"Вакансия: {vacancy.title}")
        stats_title.font = Font(bold=True, size=14, color="1F2937")
        stats_title.alignment = Alignment(horizontal="center", vertical="center")
        stats_title.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")

        # Подсчет статистики
        total_count = len(results)
        hire_count = len([r for r, _ in results if r.recommendation == 'hire'])
        interview_count = len([r for r, _ in results if r.recommendation == 'interview'])
        maybe_count = len([r for r, _ in results if r.recommendation == 'maybe'])
        reject_count = len([r for r, _ in results if r.recommendation == 'reject'])
        avg_score = sum([r.score for r, _ in results if r.score]) / total_count if total_count > 0 else 0

        # Карточки статистики (начиная со строки 6)
        stats_data = [
            ("Всего кандидатов", total_count, "4F46E5"),
            ("Средний балл", f"{avg_score:.1f}", "10B981"),
            ("Нанять ✅", hire_count, "059669"),
            ("Собеседование 👤", interview_count, "D97706"),
            ("Возможно 🤔", maybe_count, "DB2777"),
            ("Отклонить ❌", reject_count, "6B7280"),
        ]

        row = 6
        for label, value, color in stats_data:
            stats_ws.merge_cells(f'A{row}:E{row}')
            cell = stats_ws.cell(row=row, column=1, value=f"{label}: {value}")
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            stats_ws.row_dimensions[row].height = 30
            row += 1

        # Автоширина для статистики
        for col in range(1, 6):
            stats_ws.column_dimensions[chr(64 + col)].width = 30

        # Сохранение в временный файл
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        # Возврат файла
        from urllib.parse import quote

        # Безопасное имя файла (только ASCII)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_filename = f"timly_analysis_{timestamp}.xlsx"

        # URL-encoded имя для заголовка с кириллицей
        display_filename = f"timly_analysis_{vacancy.title}_{timestamp}.xlsx"
        encoded_filename = quote(display_filename)

        return FileResponse(
            temp_file.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=safe_filename,
            headers={
                "Content-Disposition": f"attachment; filename={safe_filename}; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Ошибка экспорта в Excel: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"error": "EXPORT_ERROR", "message": f"Ошибка экспорта: {str(e)}"}
        )


@router.get("/export/{export_job_id}/status", response_model=APIResponse)
async def get_export_status(
    export_job_id: str,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Проверка статуса экспорта
    Готовность файла для скачивания
    """
    # TODO: Реализовать проверку статуса экспорта
    export_status_data = {
        "export_job_id": export_job_id,
        "status": "completed",
        "file_size_bytes": 15360,
        "records_count": 25,
        "download_url": f"/api/analysis/export/{export_job_id}/download",
        "expires_at": "2024-01-03T00:00:00"
    }

    return success(data=export_status_data)


@router.get("/export/{export_job_id}/download")
async def download_export_file(
    export_job_id: str,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Скачивание экспортированного Excel файла
    Возвращает файл для скачивания
    """
    # TODO: Реализовать скачивание файла
    # file_path = f"/tmp/exports/{export_job_id}.xlsx"
    # return FileResponse(
    #     file_path,
    #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #     filename=f"timly_analysis_{export_job_id}.xlsx"
    # )

    raise HTTPException(
        status_code=status.HTTP_501_NOT_IMPLEMENTED,
        detail={
            "error": "NOT_IMPLEMENTED",
            "message": "Скачивание файлов будет реализовано"
        }
    )


@router.get("/job/{job_id}/status", response_model=APIResponse)
async def get_analysis_job_status(
    job_id: str,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Проверка статуса фоновой задачи анализа
    Прогресс выполнения и результаты
    """
    # TODO: Реализовать получение статуса задачи
    job_status_data = {
        "job_id": job_id,
        "status": "processing",
        "progress": 60,
        "total_applications": 5,
        "processed_applications": 3,
        "successful_analyses": 3,
        "failed_analyses": 0,
        "estimated_completion": "2024-01-01T10:02:00",
        "errors": []
    }

    return success(data=job_status_data)


@router.delete("/job/{job_id}", response_model=APIResponse)
async def cancel_analysis_job(
    job_id: str,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Отмена выполняющейся задачи анализа
    Останавливает фоновую обработку
    """
    # TODO: Реализовать отмену задачи
    return success(data={
        "status": "TODO",
        "message": f"Отмена задачи анализа {job_id} будет реализована",
        "job_id": job_id,
        "cancelled": True
    })


@router.get("/dashboard", response_model=APIResponse)
async def get_analysis_dashboard(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Dashboard с общей статистикой анализов
    Метрики для главной страницы с реальными данными
    """
    try:
        from app.models.application import Application, AnalysisResult
        from app.models.vacancy import Vacancy
        from sqlalchemy import func
        from datetime import datetime, timedelta

        # Общее количество анализов
        total_analyses = db.query(func.count(AnalysisResult.id)).join(
            Application
        ).join(
            Vacancy
        ).filter(
            Vacancy.user_id == current_user.id
        ).scalar() or 0

        # Анализы за этот месяц
        month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        analyses_this_month = db.query(func.count(AnalysisResult.id)).join(
            Application
        ).join(
            Vacancy
        ).filter(
            Vacancy.user_id == current_user.id,
            AnalysisResult.created_at >= month_start
        ).scalar() or 0

        # Средний балл
        avg_score = db.query(func.avg(AnalysisResult.score)).join(
            Application
        ).join(
            Vacancy
        ).filter(
            Vacancy.user_id == current_user.id,
            AnalysisResult.score.isnot(None)
        ).scalar() or 0

        # Количество топ кандидатов (score >= 80)
        top_candidates_count = db.query(func.count(AnalysisResult.id)).join(
            Application
        ).join(
            Vacancy
        ).filter(
            Vacancy.user_id == current_user.id,
            AnalysisResult.score >= 80
        ).scalar() or 0

        # Общая стоимость анализов (в рублях, конвертируем в копейки для совместимости)
        total_cost_rub = db.query(func.sum(AnalysisResult.ai_cost_rub)).join(
            Application
        ).join(
            Vacancy
        ).filter(
            Vacancy.user_id == current_user.id,
            AnalysisResult.ai_cost_rub.isnot(None)
        ).scalar() or 0
        total_cost_cents = int(float(total_cost_rub) * 100) if total_cost_rub else 0

        # Стоимость за этот месяц
        cost_this_month_rub = db.query(func.sum(AnalysisResult.ai_cost_rub)).join(
            Application
        ).join(
            Vacancy
        ).filter(
            Vacancy.user_id == current_user.id,
            AnalysisResult.created_at >= month_start,
            AnalysisResult.ai_cost_rub.isnot(None)
        ).scalar() or 0
        cost_this_month_cents = int(float(cost_this_month_rub) * 100) if cost_this_month_rub else 0

        # Недавние анализы (последние 10)
        recent_analyses_raw = db.query(
            AnalysisResult, Application, Vacancy
        ).join(
            Application,
            AnalysisResult.application_id == Application.id
        ).join(
            Vacancy,
            Application.vacancy_id == Vacancy.id
        ).filter(
            Vacancy.user_id == current_user.id
        ).order_by(
            AnalysisResult.created_at.desc()
        ).limit(10).all()

        recent_analyses = [
            {
                "vacancy_title": vacancy.title,
                "candidate_name": application.candidate_name or "Без имени",
                "score": analysis.score or 0,
                "recommendation": analysis.recommendation or "unknown",
                "analyzed_at": analysis.created_at.isoformat() if analysis.created_at else None
            }
            for analysis, application, vacancy in recent_analyses_raw
        ]

        dashboard_data = {
            "total_analyses": total_analyses,
            "analyses_this_month": analyses_this_month,
            "avg_score": round(float(avg_score), 1) if avg_score else 0,
            "top_candidates_count": top_candidates_count,
            "total_cost_cents": total_cost_cents,
            "cost_this_month_cents": cost_this_month_cents,
            "recent_analyses": recent_analyses
        }

        return success(data=dashboard_data)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Ошибка получения статистики dashboard: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"error": "DASHBOARD_ERROR", "message": f"Ошибка получения статистики: {str(e)}"}
        )