"""
API для ручного анализа резюме без привязки к HH.ru
Загрузка файлов, создание вакансий вручную, AI-анализ
"""
import logging
import io
import tempfile
import os
from typing import List, Optional, Dict, Any
from datetime import datetime
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, Query, Body
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.database import get_db
from app.models.user import User
from app.api.auth import get_current_user
from app.services.resume_parser import ResumeParser
from app.services.ai_analyzer import AIAnalyzer
from app.services.subscription_service import SubscriptionService
from app.utils.exceptions import FileParseError, AIAnalysisError

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/manual-analysis", tags=["Manual Analysis"])


# ==================== SCHEMAS ====================

class ManualVacancy(BaseModel):
    """Схема вакансии для ручного анализа"""
    title: str
    description: Optional[str] = ""
    key_skills: List[str] = []
    experience_required: str = "1-3"
    salary_from: Optional[int] = None
    salary_to: Optional[int] = None
    currency: str = "RUB"


class ParsedCandidate(BaseModel):
    """Схема распарсенного кандидата"""
    id: str
    full_name: str
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    title: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    city: Optional[str] = None
    experience_years: Optional[int] = None
    skills: List[str] = []
    salary_expectation: Optional[int] = None
    experience_text: Optional[str] = None
    education: Optional[str] = None
    original_text: Optional[str] = None


class AnalyzeRequest(BaseModel):
    """Запрос на анализ кандидата"""
    vacancy: ManualVacancy
    candidate: ParsedCandidate


class ExportRequest(BaseModel):
    """Запрос на экспорт в Excel"""
    vacancy: ManualVacancy
    results: List[Dict[str, Any]]


# ==================== ENDPOINTS ====================

@router.post("/parse/pdf")
async def parse_pdf_resume(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Парсинг PDF резюме

    - Извлекает текст из PDF
    - AI-извлечение структурированных данных
    - Возвращает данные кандидата
    """
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате PDF")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл слишком большой (макс 10 MB)")

    try:
        parser = ResumeParser()
        parsed_data = await parser.parse_pdf(io.BytesIO(content), file.filename)

        # Формируем ответ
        candidate_id = str(uuid4())
        full_name = f"{parsed_data.get('first_name', '')} {parsed_data.get('last_name', '')}".strip()
        if not full_name:
            full_name = parsed_data.get('full_name', 'Без имени')

        return ParsedCandidate(
            id=candidate_id,
            full_name=full_name,
            first_name=parsed_data.get('first_name'),
            last_name=parsed_data.get('last_name'),
            title=parsed_data.get('title'),
            email=parsed_data.get('email'),
            phone=parsed_data.get('phone'),
            city=parsed_data.get('city'),
            experience_years=parsed_data.get('experience_years'),
            skills=parsed_data.get('skills', []),
            salary_expectation=parsed_data.get('salary_expectation'),
            experience_text=parsed_data.get('experience_text'),
            education=parsed_data.get('education'),
            original_text=parsed_data.get('original_text')
        )

    except FileParseError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Ошибка парсинга PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка обработки файла: {str(e)}")


@router.post("/parse/document")
async def parse_document_resume(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Парсинг DOCX резюме

    - Извлекает текст из DOCX
    - AI-извлечение структурированных данных
    """
    if not file.filename.lower().endswith(('.docx', '.doc')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате DOCX/DOC")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл слишком большой (макс 10 MB)")

    try:
        parser = ResumeParser()
        parsed_data = await parser.parse_docx(io.BytesIO(content), file.filename)

        candidate_id = str(uuid4())
        full_name = f"{parsed_data.get('first_name', '')} {parsed_data.get('last_name', '')}".strip()
        if not full_name:
            full_name = parsed_data.get('full_name', 'Без имени')

        return ParsedCandidate(
            id=candidate_id,
            full_name=full_name,
            first_name=parsed_data.get('first_name'),
            last_name=parsed_data.get('last_name'),
            title=parsed_data.get('title'),
            email=parsed_data.get('email'),
            phone=parsed_data.get('phone'),
            city=parsed_data.get('city'),
            experience_years=parsed_data.get('experience_years'),
            skills=parsed_data.get('skills', []),
            salary_expectation=parsed_data.get('salary_expectation'),
            experience_text=parsed_data.get('experience_text'),
            education=parsed_data.get('education'),
            original_text=parsed_data.get('original_text')
        )

    except FileParseError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Ошибка парсинга DOCX: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка обработки файла: {str(e)}")


@router.post("/parse/excel")
async def parse_excel_resumes(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Парсинг Excel файла с резюме

    - Автоопределение колонок
    - Возвращает список кандидатов
    """
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате Excel")

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл слишком большой (макс 50 MB)")

    try:
        parser = ResumeParser()
        candidates_data = await parser.parse_excel(io.BytesIO(content), file.filename)

        candidates = []
        for data in candidates_data:
            candidate_id = str(uuid4())
            full_name = f"{data.get('first_name', '')} {data.get('last_name', '')}".strip()
            if not full_name:
                full_name = data.get('full_name', 'Без имени')

            candidates.append(ParsedCandidate(
                id=candidate_id,
                full_name=full_name,
                first_name=data.get('first_name'),
                last_name=data.get('last_name'),
                title=data.get('title'),
                email=data.get('email'),
                phone=data.get('phone'),
                city=data.get('city'),
                experience_years=data.get('experience_years'),
                skills=data.get('skills', []),
                salary_expectation=data.get('salary_expectation')
            ))

        return {"candidates": candidates, "count": len(candidates)}

    except FileParseError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Ошибка парсинга Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка обработки файла: {str(e)}")


@router.post("/analyze")
async def analyze_candidate(
    request: AnalyzeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    AI-анализ кандидата под вакансию

    - Проверяет лимиты подписки
    - Анализирует резюме через AI
    - Возвращает детальный результат
    """
    # Проверка лимитов
    subscription_service = SubscriptionService(db)
    can_analyze, error_msg = await subscription_service.check_can_analyze(current_user.id)
    if not can_analyze:
        subscription = await subscription_service.get_or_create_subscription(current_user.id)
        raise HTTPException(
            status_code=403,
            detail={
                "error": "LIMIT_EXCEEDED",
                "message": error_msg or "Лимит анализов исчерпан",
                "subscription": subscription.to_dict()
            }
        )

    try:
        # Формируем данные вакансии
        vacancy_data = {
            "title": request.vacancy.title,
            "description": request.vacancy.description,
            "key_skills": request.vacancy.key_skills,
            "experience": request.vacancy.experience_required,
            "salary_from": request.vacancy.salary_from,
            "salary_to": request.vacancy.salary_to,
            "currency": request.vacancy.currency
        }

        # Формируем данные резюме в формате HH-like
        resume_data = {
            "first_name": request.candidate.first_name,
            "last_name": request.candidate.last_name,
            "title": request.candidate.title,
            "age": None,
            "area": {"name": request.candidate.city} if request.candidate.city else None,
            "total_experience": {"months": (request.candidate.experience_years or 0) * 12},
            "salary": {"amount": request.candidate.salary_expectation} if request.candidate.salary_expectation else None,
            "skill_set": [{"name": skill} for skill in request.candidate.skills],
            "experience": [
                {
                    "company": {"name": "Указан в резюме"},
                    "position": request.candidate.title or "Не указано",
                    "description": request.candidate.experience_text or ""
                }
            ] if request.candidate.experience_text else []
        }

        # AI анализ
        analyzer = AIAnalyzer()
        analysis = await analyzer.analyze_resume(vacancy_data, resume_data)

        # Увеличиваем счетчик использования
        await subscription_service.increment_analysis_usage(current_user.id)

        return {
            "id": request.candidate.id,
            "score": analysis.get("score", 50),
            "recommendation": analysis.get("recommendation", "maybe"),
            "skills_match": analysis.get("skills_match", 50),
            "experience_match": analysis.get("experience_match", 50),
            "salary_match": analysis.get("salary_match", "unknown"),
            "career_trajectory": analysis.get("career_trajectory", "unknown"),
            "skill_gaps": analysis.get("skill_gaps", []),
            "strengths": analysis.get("strengths", []),
            "weaknesses": analysis.get("weaknesses", []),
            "red_flags": analysis.get("red_flags", []),
            "green_flags": analysis.get("green_flags", []),
            "reasoning": analysis.get("reasoning", ""),
            "interview_questions": analysis.get("interview_questions", [])
        }

    except AIAnalysisError as e:
        raise HTTPException(status_code=500, detail=f"Ошибка AI анализа: {str(e)}")
    except Exception as e:
        logger.error(f"Ошибка анализа кандидата: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка анализа: {str(e)}")


@router.post("/export/excel")
async def export_to_excel(
    request: ExportRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Экспорт результатов анализа в Excel

    - Формирует структурированный Excel файл
    - Включает все результаты анализа
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ==================== ЛИСТ 1: СВОДКА ====================
        ws = wb.active
        ws.title = "Сводка"

        # Стили
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='333333'),
            right=Side(style='thin', color='333333'),
            top=Side(style='thin', color='333333'),
            bottom=Side(style='thin', color='333333')
        )

        # Заголовки
        headers = ["№", "ФИО", "Должность", "Балл", "Рекомендация", "Навыки %", "Опыт %", "Карьера"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Данные
        for idx, result in enumerate(request.results, 1):
            candidate = result.get("candidate", {})
            analysis = result.get("analysis", {})

            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=candidate.get("full_name", "—"))
            ws.cell(row=row, column=3, value=candidate.get("title", "—"))
            ws.cell(row=row, column=4, value=analysis.get("score", "—"))
            ws.cell(row=row, column=5, value=_get_recommendation_text(analysis.get("recommendation")))
            ws.cell(row=row, column=6, value=f"{analysis.get('skills_match', 0)}%")
            ws.cell(row=row, column=7, value=f"{analysis.get('experience_match', 0)}%")
            ws.cell(row=row, column=8, value=_get_career_text(analysis.get("career_trajectory")))

            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border

        # Ширина колонок
        col_widths = [5, 25, 20, 8, 15, 10, 10, 12]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # ==================== ЛИСТ 2: ДЕТАЛИ ====================
        ws2 = wb.create_sheet("Детальный анализ")

        headers2 = ["ФИО", "Должность", "Контакты", "Балл", "Рекомендация",
                    "Сильные стороны", "Слабые стороны", "Недостающие навыки",
                    "Вопросы для интервью", "Обоснование AI"]

        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for idx, result in enumerate(request.results, 1):
            candidate = result.get("candidate", {})
            analysis = result.get("analysis", {})

            row = idx + 1
            contacts = []
            if candidate.get("email"):
                contacts.append(candidate["email"])
            if candidate.get("phone"):
                contacts.append(candidate["phone"])

            ws2.cell(row=row, column=1, value=candidate.get("full_name", "—"))
            ws2.cell(row=row, column=2, value=candidate.get("title", "—"))
            ws2.cell(row=row, column=3, value="\n".join(contacts) if contacts else "—")
            ws2.cell(row=row, column=4, value=analysis.get("score", "—"))
            ws2.cell(row=row, column=5, value=_get_recommendation_text(analysis.get("recommendation")))
            ws2.cell(row=row, column=6, value=_format_list(analysis.get("strengths", [])))
            ws2.cell(row=row, column=7, value=_format_list(analysis.get("weaknesses", [])))
            ws2.cell(row=row, column=8, value=_format_list(analysis.get("skill_gaps", [])))
            ws2.cell(row=row, column=9, value=_format_list(analysis.get("interview_questions", [])))
            ws2.cell(row=row, column=10, value=analysis.get("reasoning", "—"))

            for col in range(1, 11):
                ws2.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

        # Ширина колонок
        col_widths2 = [20, 18, 20, 8, 12, 30, 30, 25, 35, 40]
        for col, width in enumerate(col_widths2, 1):
            ws2.column_dimensions[get_column_letter(col)].width = width

        # ==================== ЛИСТ 3: СТАТИСТИКА ====================
        ws3 = wb.create_sheet("Статистика")

        ws3.cell(row=1, column=1, value="Вакансия")
        ws3.cell(row=1, column=2, value=request.vacancy.title)
        ws3.cell(row=2, column=1, value="Всего кандидатов")
        ws3.cell(row=2, column=2, value=len(request.results))

        # Распределение по рекомендациям
        recs = {"hire": 0, "interview": 0, "maybe": 0, "reject": 0}
        total_score = 0
        for r in request.results:
            rec = r.get("analysis", {}).get("recommendation")
            if rec in recs:
                recs[rec] += 1
            total_score += r.get("analysis", {}).get("score", 0)

        ws3.cell(row=4, column=1, value="Нанять")
        ws3.cell(row=4, column=2, value=recs["hire"])
        ws3.cell(row=5, column=1, value="Собеседование")
        ws3.cell(row=5, column=2, value=recs["interview"])
        ws3.cell(row=6, column=1, value="Возможно")
        ws3.cell(row=6, column=2, value=recs["maybe"])
        ws3.cell(row=7, column=1, value="Отклонить")
        ws3.cell(row=7, column=2, value=recs["reject"])

        ws3.cell(row=9, column=1, value="Средний балл")
        avg_score = round(total_score / len(request.results)) if request.results else 0
        ws3.cell(row=9, column=2, value=avg_score)

        # Сохранение
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"manual_analysis_{timestamp}.xlsx"

        return FileResponse(
            temp_file.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Ошибка экспорта в Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка экспорта: {str(e)}")


# ==================== HELPERS ====================

def _get_recommendation_text(rec: str) -> str:
    mapping = {
        "hire": "Нанять",
        "interview": "Собеседование",
        "maybe": "Возможно",
        "reject": "Отклонить"
    }
    return mapping.get(rec, "—")


def _get_career_text(career: str) -> str:
    mapping = {
        "growth": "Рост",
        "stable": "Стабильно",
        "decline": "Спад",
        "unknown": "—"
    }
    return mapping.get(career, "—")


def _format_list(items: List[str]) -> str:
    if not items:
        return "—"
    return "\n".join([f"• {item}" for item in items[:5]])
