#!/usr/bin/env python3
"""
Генерация примера Excel экспорта v2.0 с тестовыми данными
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Тестовые данные кандидатов
test_candidates = [
    {
        "name": "Иванов Алексей Петрович",
        "email": "ivanov@mail.ru",
        "phone": "+7 (999) 123-45-67",
        "resume_url": "https://hh.ru/resume/123456",
        "score": 92,
        "skills_match": 95,
        "experience_match": 88,
        "career_trajectory": "growth",
        "recommendation": "hire",
        "salary_match": "match",
        "reasoning": "Отличный кандидат с релевантным опытом в разработке на Python. 5 лет в IT, последние 2 года - Senior Developer.",
        "strengths": ["5+ лет опыта Python/FastAPI", "Опыт работы с высоконагруженными системами", "Хорошие soft skills по отзывам"],
        "weaknesses": ["Нет опыта с Kubernetes", "Не работал с микросервисами"],
        "skill_gaps": ["Kubernetes", "Docker Swarm"],
        "interview_questions": ["Расскажите о самом сложном техническом решении", "Как бы вы спроектировали систему уведомлений?"],
        "red_flags": [],
    },
    {
        "name": "Петрова Мария Сергеевна",
        "email": "petrova.m@gmail.com",
        "phone": "+7 (925) 987-65-43",
        "resume_url": "https://hh.ru/resume/234567",
        "score": 78,
        "skills_match": 82,
        "experience_match": 75,
        "career_trajectory": "stable",
        "recommendation": "interview",
        "salary_match": "higher",
        "reasoning": "Хороший кандидат, но зарплатные ожидания выше бюджета. Стоит обсудить на собеседовании.",
        "strengths": ["Сильный бэкграунд в аналитике", "Опыт работы с ML", "Быстрая обучаемость"],
        "weaknesses": ["Зарплатные ожидания +30% от бюджета", "Нет опыта руководства"],
        "skill_gaps": ["Team Lead опыт", "Agile/Scrum"],
        "interview_questions": ["Готовы ли вы к пересмотру зарплатных ожиданий?", "Какой у вас опыт работы в команде?"],
        "red_flags": [],
    },
    {
        "name": "Сидоров Дмитрий Владимирович",
        "email": "sidorov.d@yandex.ru",
        "phone": "+7 (916) 555-44-33",
        "resume_url": "https://hh.ru/resume/345678",
        "score": 65,
        "skills_match": 60,
        "experience_match": 70,
        "career_trajectory": "stable",
        "recommendation": "maybe",
        "salary_match": "lower",
        "reasoning": "Средний кандидат. Не хватает ключевых навыков, но есть потенциал для роста.",
        "strengths": ["Мотивирован к обучению", "Хорошее техническое образование", "Адекватные ожидания по зарплате"],
        "weaknesses": ["Только 2 года опыта", "Не работал с нашим стеком", "Нет коммерческого опыта с PostgreSQL"],
        "skill_gaps": ["FastAPI", "PostgreSQL", "Redis", "Celery"],
        "interview_questions": ["Как быстро вы сможете освоить наш стек?", "Расскажите о своём опыте самообучения"],
        "red_flags": ["Частая смена работы (3 места за 2 года)"],
    },
    {
        "name": "Козлова Анна Игоревна",
        "email": "kozlova.anna@inbox.ru",
        "phone": "+7 (903) 222-11-00",
        "resume_url": "https://hh.ru/resume/456789",
        "score": 45,
        "skills_match": 40,
        "experience_match": 50,
        "career_trajectory": "decline",
        "recommendation": "reject",
        "salary_match": "unknown",
        "reasoning": "Не подходит по ключевым требованиям. Последняя позиция ниже предыдущей.",
        "strengths": ["Опыт работы в крупной компании"],
        "weaknesses": ["Downgrade с Senior на Middle", "Пробел в резюме 8 месяцев", "Нет 70% требуемых навыков"],
        "skill_gaps": ["Python", "FastAPI", "PostgreSQL", "Docker", "CI/CD", "Git"],
        "interview_questions": [],
        "red_flags": ["Понижение в должности", "Пробел в резюме >6 месяцев", "Отсутствует большинство ключевых навыков"],
    },
    {
        "name": "Новиков Артём Александрович",
        "email": "novikov.art@mail.ru",
        "phone": "+7 (977) 888-77-66",
        "resume_url": "https://hh.ru/resume/567890",
        "score": 85,
        "skills_match": 90,
        "experience_match": 80,
        "career_trajectory": "growth",
        "recommendation": "interview",
        "salary_match": "match",
        "reasoning": "Сильный кандидат с хорошим ростом. Рекомендуется техническое собеседование.",
        "strengths": ["Быстрый карьерный рост за 3 года", "Полный стек технологий", "Опыт менторинга джунов"],
        "weaknesses": ["Нет опыта с высокими нагрузками", "Работал только в стартапах"],
        "skill_gaps": ["Highload архитектура"],
        "interview_questions": ["Как бы вы подошли к оптимизации медленных запросов?", "Расскажите о вашем опыте менторинга"],
        "red_flags": [],
    },
]

def format_list_as_bullets(items):
    """Форматирует список в строку с буллетами"""
    if not items:
        return "—"
    return "\n".join([f"• {item}" for item in items if item])

def create_example_excel():
    """Создаёт пример Excel файла"""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary_ws = wb.create_sheet("Сводка", 0)
    details_ws = wb.create_sheet("Детальный анализ", 1)
    stats_ws = wb.create_sheet("Статистика", 2)

    # ========== СТИЛИ ==========
    BRAND_PRIMARY = "6366F1"
    HEADER_BG = "4F46E5"
    LIGHT_BG = "F3F4F6"

    title_font = Font(bold=True, size=18, color="FFFFFF", name="Arial")
    header_font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
    subheader_font = Font(bold=True, size=10, color="374151", name="Arial")
    normal_font = Font(size=10, color="374151", name="Arial")
    link_font = Font(size=10, color="2563EB", underline="single")

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    brand_fill = PatternFill(start_color=BRAND_PRIMARY, end_color=BRAND_PRIMARY, fill_type="solid")
    light_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

    rec_styles = {
        'hire': {'fill': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"), 'font': Font(bold=True, size=10, color="065F46"), 'text': "НАНЯТЬ"},
        'interview': {'fill': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"), 'font': Font(bold=True, size=10, color="92400E"), 'text': "СОБЕСЕДОВАНИЕ"},
        'maybe': {'fill': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), 'font': Font(bold=True, size=10, color="991B1B"), 'text': "ВОЗМОЖНО"},
        'reject': {'fill': PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"), 'font': Font(bold=True, size=10, color="6B7280"), 'text': "ОТКЛОНИТЬ"}
    }

    career_styles = {
        'growth': {'fill': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"), 'font': Font(bold=True, size=10, color="065F46"), 'text': "📈 Рост"},
        'stable': {'fill': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"), 'font': Font(bold=True, size=10, color="92400E"), 'text': "➡️ Стабильно"},
        'decline': {'fill': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), 'font': Font(bold=True, size=10, color="991B1B"), 'text': "📉 Снижение"},
        'unknown': {'fill': PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"), 'font': Font(size=10, color="6B7280"), 'text': "❓ Н/Д"}
    }

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    vacancy_title = "Senior Python Developer"

    # ========== ЛИСТ 1: СВОДКА ==========
    summary_ws.merge_cells('A1:K2')
    title_cell = summary_ws.cell(row=1, column=1, value=f"TIMLY | Анализ кандидатов: {vacancy_title}")
    title_cell.font = title_font
    title_cell.fill = brand_fill
    title_cell.alignment = center_align
    summary_ws.row_dimensions[1].height = 25
    summary_ws.row_dimensions[2].height = 25

    summary_ws.merge_cells('A3:K3')
    date_cell = summary_ws.cell(row=3, column=1, value=f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Всего кандидатов: {len(test_candidates)}")
    date_cell.font = Font(size=10, color="6B7280", italic=True)
    date_cell.alignment = center_align
    date_cell.fill = light_fill

    # Заголовки с фиксированными ширинами
    summary_headers = [
        ("№", 5), ("Кандидат", 28), ("Контакты", 28), ("Оценка", 10),
        ("Навыки %", 10), ("Опыт %", 10), ("Карьера", 14),
        ("Рекомендация", 16), ("Зарплата", 14), ("Ключевое", 50), ("Резюме", 12)
    ]

    for col, (header, width) in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        summary_ws.column_dimensions[get_column_letter(col)].width = width

    summary_ws.row_dimensions[5].height = 25
    summary_ws.freeze_panes = "A6"

    # Данные сводки
    for idx, candidate in enumerate(test_candidates, 1):
        row = idx + 5

        summary_ws.cell(row=row, column=1, value=idx).alignment = center_align
        summary_ws.cell(row=row, column=2, value=candidate["name"]).font = subheader_font

        # Контакты
        contacts = f"{candidate['email']}\n{candidate['phone']}"
        summary_ws.cell(row=row, column=3, value=contacts).alignment = left_align

        # Оценка
        score_cell = summary_ws.cell(row=row, column=4, value=candidate["score"])
        score_cell.alignment = center_align
        score_cell.font = Font(bold=True, size=12)

        # Навыки и опыт
        summary_ws.cell(row=row, column=5, value=candidate["skills_match"]).alignment = center_align
        summary_ws.cell(row=row, column=6, value=candidate["experience_match"]).alignment = center_align

        # Карьерная траектория
        career = candidate.get("career_trajectory", "unknown")
        career_style = career_styles.get(career, career_styles['unknown'])
        career_cell = summary_ws.cell(row=row, column=7, value=career_style['text'])
        career_cell.fill = career_style['fill']
        career_cell.font = career_style['font']
        career_cell.alignment = center_align

        # Рекомендация
        rec = candidate["recommendation"]
        rec_style = rec_styles.get(rec, rec_styles['maybe'])
        rec_cell = summary_ws.cell(row=row, column=8, value=rec_style['text'])
        rec_cell.fill = rec_style['fill']
        rec_cell.font = rec_style['font']
        rec_cell.alignment = center_align

        # Зарплата
        salary_map = {'match': '✅ Совпадает', 'higher': '⬆️ Выше', 'lower': '⬇️ Ниже', 'unknown': '❓ Н/Д'}
        summary_ws.cell(row=row, column=9, value=salary_map.get(candidate["salary_match"], '❓ Н/Д')).alignment = center_align

        # Ключевое обоснование
        reasoning_cell = summary_ws.cell(row=row, column=10, value=candidate["reasoning"])
        reasoning_cell.alignment = left_align
        reasoning_cell.font = normal_font

        # Ссылка на резюме
        link_cell = summary_ws.cell(row=row, column=11, value="Открыть →")
        link_cell.hyperlink = candidate["resume_url"]
        link_cell.font = link_font
        link_cell.alignment = center_align

        # Границы
        for col in range(1, 12):
            summary_ws.cell(row=row, column=col).border = thin_border
        summary_ws.row_dimensions[row].height = 50

    # Условное форматирование для оценок
    score_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FEE2E2',
        mid_type='num', mid_value=60, mid_color='FEF3C7',
        end_type='num', end_value=100, end_color='D1FAE5'
    )
    summary_ws.conditional_formatting.add(f"D6:D{len(test_candidates) + 5}", score_rule)
    summary_ws.conditional_formatting.add(f"E6:E{len(test_candidates) + 5}", score_rule)
    summary_ws.conditional_formatting.add(f"F6:F{len(test_candidates) + 5}", score_rule)
    summary_ws.auto_filter.ref = f"A5:K{len(test_candidates) + 5}"

    # ========== ЛИСТ 2: ДЕТАЛЬНЫЙ АНАЛИЗ ==========
    details_ws.merge_cells('A1:H2')
    title_cell = details_ws.cell(row=1, column=1, value=f"TIMLY | Детальный анализ: {vacancy_title}")
    title_cell.font = title_font
    title_cell.fill = brand_fill
    title_cell.alignment = center_align
    details_ws.row_dimensions[1].height = 25
    details_ws.row_dimensions[2].height = 25

    # Ширины колонок
    detail_widths = [5, 28, 12, 16, 40, 40, 35, 40]
    for col, width in enumerate(detail_widths, 1):
        details_ws.column_dimensions[get_column_letter(col)].width = width

    detail_headers = ["№", "Кандидат", "Оценка", "Рекомендация", "Сильные стороны", "Слабые стороны", "Skill Gaps", "Вопросы для интервью"]
    for col, header in enumerate(detail_headers, 1):
        cell = details_ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    details_ws.row_dimensions[4].height = 25
    details_ws.freeze_panes = "A5"

    for idx, candidate in enumerate(test_candidates, 1):
        row = idx + 4

        details_ws.cell(row=row, column=1, value=idx).alignment = center_align
        name_cell = details_ws.cell(row=row, column=2, value=candidate["name"])
        name_cell.font = subheader_font
        name_cell.alignment = left_align

        score_cell = details_ws.cell(row=row, column=3, value=candidate["score"])
        score_cell.alignment = center_align
        score_cell.font = Font(bold=True, size=14)

        rec = candidate["recommendation"]
        rec_style = rec_styles.get(rec, rec_styles['maybe'])
        rec_cell = details_ws.cell(row=row, column=4, value=rec_style['text'])
        rec_cell.fill = rec_style['fill']
        rec_cell.font = rec_style['font']
        rec_cell.alignment = center_align

        # Сильные стороны с буллетами
        strengths_cell = details_ws.cell(row=row, column=5, value=format_list_as_bullets(candidate["strengths"]))
        strengths_cell.alignment = left_align
        strengths_cell.font = normal_font

        # Слабые стороны с буллетами
        weaknesses_cell = details_ws.cell(row=row, column=6, value=format_list_as_bullets(candidate["weaknesses"]))
        weaknesses_cell.alignment = left_align
        weaknesses_cell.font = normal_font

        # Skill Gaps
        skill_gaps_cell = details_ws.cell(row=row, column=7, value=format_list_as_bullets(candidate["skill_gaps"]))
        skill_gaps_cell.alignment = left_align
        skill_gaps_cell.font = normal_font

        # Вопросы для интервью
        questions_cell = details_ws.cell(row=row, column=8, value=format_list_as_bullets(candidate["interview_questions"]))
        questions_cell.alignment = left_align
        questions_cell.font = normal_font

        for col in range(1, 9):
            details_ws.cell(row=row, column=col).border = thin_border
        details_ws.row_dimensions[row].height = 90

    details_ws.conditional_formatting.add(f"C5:C{len(test_candidates) + 4}", score_rule)
    details_ws.auto_filter.ref = f"A4:H{len(test_candidates) + 4}"

    # ========== ЛИСТ 3: СТАТИСТИКА ==========
    stats_ws.merge_cells('A1:D2')
    title_cell = stats_ws.cell(row=1, column=1, value="TIMLY | Статистика анализа")
    title_cell.font = title_font
    title_cell.fill = brand_fill
    title_cell.alignment = center_align
    stats_ws.row_dimensions[1].height = 25
    stats_ws.row_dimensions[2].height = 25

    stats_ws.merge_cells('A3:D3')
    vacancy_cell = stats_ws.cell(row=3, column=1, value=f"Вакансия: {vacancy_title}")
    vacancy_cell.font = subheader_font
    vacancy_cell.alignment = center_align
    vacancy_cell.fill = light_fill

    for col in range(1, 5):
        stats_ws.column_dimensions[get_column_letter(col)].width = 25

    # Подсчёт статистики
    total_count = len(test_candidates)
    hire_count = len([c for c in test_candidates if c["recommendation"] == 'hire'])
    interview_count = len([c for c in test_candidates if c["recommendation"] == 'interview'])
    maybe_count = len([c for c in test_candidates if c["recommendation"] == 'maybe'])
    reject_count = len([c for c in test_candidates if c["recommendation"] == 'reject'])

    scores = [c["score"] for c in test_candidates]
    avg_score = sum(scores) / len(scores)
    max_score = max(scores)
    min_score_val = min(scores)

    growth_count = len([c for c in test_candidates if c.get("career_trajectory") == 'growth'])
    stable_count = len([c for c in test_candidates if c.get("career_trajectory") == 'stable'])
    decline_count = len([c for c in test_candidates if c.get("career_trajectory") == 'decline'])

    stats_data = [
        ("📊 ОБЩИЕ МЕТРИКИ", "", "", ""),
        ("Всего кандидатов", total_count, "Средний балл", f"{avg_score:.1f}"),
        ("Макс. балл", max_score, "Мин. балл", min_score_val),
        ("", "", "", ""),
        ("📋 РЕКОМЕНДАЦИИ", "", "", ""),
        ("✅ Нанять", hire_count, "👤 Собеседование", interview_count),
        ("🤔 Возможно", maybe_count, "❌ Отклонить", reject_count),
        ("", "", "", ""),
        ("📈 КАРЬЕРНАЯ ДИНАМИКА", "", "", ""),
        ("Рост", growth_count, "Стабильно", stable_count),
        ("Снижение", decline_count, "", ""),
    ]

    row = 5
    for data in stats_data:
        for col, val in enumerate(data, 1):
            cell = stats_ws.cell(row=row, column=col, value=val)
            if col in [1, 3]:
                cell.font = subheader_font
            else:
                cell.font = Font(bold=True, size=12, color="4F46E5")
            cell.alignment = center_align

            if str(val).startswith("📊") or str(val).startswith("📋") or str(val).startswith("📈"):
                stats_ws.merge_cells(f'A{row}:D{row}')
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = brand_fill
                break
        stats_ws.row_dimensions[row].height = 28
        row += 1

    # Сохранение
    output_path = os.path.join(os.path.dirname(__file__), "timly_example_export.xlsx")
    wb.save(output_path)
    print(f"✅ Пример Excel файла создан: {output_path}")
    return output_path

if __name__ == "__main__":
    create_example_excel()
